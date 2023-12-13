import csv
import pandas as pd
from dateutil.parser import parse
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

def process_first_csv(filename, ws):
    header_to_use = "DNS Hostname" if "Domain_Server" in filename else "NetBIOS Hostname"
    hostname_control_map = defaultdict(list)

    with open(filename, 'r', encoding='utf-8-sig') as file:
        reader = csv.reader(file)
        headers = next(reader)
        dns_hostname_idx = headers.index(header_to_use)
        control_idx = headers.index("Control")

        for row in reader:
            dns_hostname = row[dns_hostname_idx]
            control = row[control_idx]
            hostname_control_map[dns_hostname].append(control)

    summary_data = [["Hostname", "Count of Control"]]
    grand_total = 0

    # Create a list of tuples [(hostname, count), ...] and sort it based on count
    sorted_hostnames = sorted(hostname_control_map.items(), key=lambda x: len(set(x[1])), reverse=True)

    for dns_hostname, controls in sorted_hostnames:
        summary_data.append([dns_hostname, len(set(controls))])
        grand_total += len(set(controls))
        control_count_map = defaultdict(int)
        for control in controls:
            control_count_map[control] += 1
        for control, count in control_count_map.items():
            summary_data.append(["       " + control, count])

    summary_data.append(["Grand Total", grand_total])

    dns_hostname_fill = PatternFill(start_color="FF9bddff", end_color="FF9bddff", fill_type="solid")
    other_fill = PatternFill(start_color="FFe0ffff", end_color="FFe0ffff", fill_type="solid")

    for row_num, row_data in enumerate(summary_data, 1):
        for col_num, cell_data in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_data)
            if isinstance(cell_data, str) and ("       " not in cell_data):
                fill_color = dns_hostname_fill
            else:
                fill_color = other_fill
            cell.fill = fill_color
            if col_num == 2 and "       " not in ws.cell(row=row_num, column=col_num-1).value:
                cell.fill = dns_hostname_fill
    
    # Adjusting the width for column A (Hostname) based on its content
    max_length = 0
    for cell in ws["A"]:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    # adjusted_width = (max_length + 1)  # adding a little extra space
    ws.column_dimensions['A'].width = max_length

    # Adjusting the width for column B (Count of Control) based on its content
    max_length = 0
    for cell in ws["B"]:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    # adjusted_width = (max_length + 1)  # adding a little extra space
    ws.column_dimensions['B'].width = max_length

def process_control_csv(filename, ws):
    with open(filename, 'r', encoding='utf-8-sig') as file:        
        df = pd.read_csv(file)
        
    grouped_data = df.groupby(['Control ID', 'Control']).agg({'Host IP': 'count'}).reset_index()
    grouped_data.columns = ["Control ID", "Control", "Count of Host IP"]
    
    # Sort the grouped data by 'Count of Host IP' column in descending order
    grouped_data = grouped_data.sort_values(by='Count of Host IP', ascending=False)
    
    # Convert DataFrame to rows suitable for openpyxl
    rows = dataframe_to_rows(grouped_data, index=False, header=True)

    # Append rows to the worksheet
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
            
    # Calculate Grand Total for 'Count of Host IP'
    grand_total = grouped_data['Count of Host IP'].sum()
    
    # Append "Grand Total" label below "Control ID" column
    ws.append(["Grand Total", None, grand_total])
    
    header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    for cell in ws[1]:  # Assuming the header is in the first row
        cell.fill = header_fill
    
    # Adjusting column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length
        
def process_host_csv(filename, ws):
    with open(filename, 'r', encoding='utf-8-sig') as file:        
        data = pd.read_csv(file)
        
    # Extracting date and time elements from 'Last Scan Date' column
    data['Last Scan Date'] = data['Last Scan Date'].apply(lambda x: parse(x.split(' at')[0]))

    # Selecting necessary columns
    selected_columns = ["Host IP", "DNS Hostname", "Last Scan Date"]

    # Grouping by "Host IP" and aggregating
    grouped_data = data[selected_columns].groupby('Host IP').agg(
        {'DNS Hostname': 'first', 'Last Scan Date': ['max', 'count']}
    ).reset_index()

    # Flattening the multi-index columns
    grouped_data.columns = [' '.join(col).strip() for col in grouped_data.columns.values]

    # Renaming the columns for the summary
    grouped_data.columns = ["Host IP", "DNS Hostname", "Last Scan Date (Latest)", "Count of Host IP"]
    
    # Sort the grouped data by 'Count of Host IP' column in descending order
    grouped_data = grouped_data.sort_values(by='Count of Host IP', ascending=False)
    
    # Convert DataFrame to rows suitable for openpyxl
    rows = dataframe_to_rows(grouped_data, index=False, header=True)

    # Append rows to the worksheet
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
            
    # Calculate Grand Total for 'Count of Host IP'
    grand_total = grouped_data['Count of Host IP'].sum()
    
    # Append "Grand Total" label below "Control ID" column
    ws.append(["Grand Total", None, None, grand_total])
    
    header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    for cell in ws[1]:  # Assuming the header is in the first row
        cell.fill = header_fill
    
    # Adjusting column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length

def process_second_csv(filename, ws):
    with open(filename, 'r', encoding='utf-8-sig') as file:
        reader = csv.reader(file)
        for row in reader:
            ws.append(row)
            
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length

def generate_excel1(first_csv_file, second_csv_file, output_excel_file):
    wb = Workbook()
    
    ws1 = wb.active
    ws1.title = "First Sheet"
    process_first_csv(first_csv_file, ws1)

    ws2 = wb.create_sheet("Second Sheet")
    process_second_csv(second_csv_file, ws2)

    wb.save(output_excel_file)

def generate_excel2(first_csv_file, output_excel_file):
    wb = Workbook()
    
    ws1 = wb.active
    ws1.title = "By Control"
    process_control_csv(first_csv_file, ws1)

    ws2 = wb.create_sheet("By Host")
    process_host_csv(first_csv_file, ws2)
    
    ws3 = wb.create_sheet("Second Sheet")
    process_second_csv(first_csv_file, ws3)

    wb.save(output_excel_file)
# Call the function
generate_excel1('TLT_Policy_Compliance_Report_Windows_2019_Domain_Server_Monthly_temp2.csv', 'TLT_Policy_Compliance_Report_Windows_2019_Domain_Server_Monthly_latest.csv', 'TLT_Policy_Compliance_Report_Windows_2019_Domain_Server_Monthly_latest.xlsx')

generate_excel1('TLT_Policy_Compliance_Report_Windows_2019_Member_Server_Monthly_temp2.csv', 'TLT_Policy_Compliance_Report_Windows_2019_Member_Server_Monthly_latest.csv', 'TLT_Policy_Compliance_Report_Windows_2019_Member_Server_Monthly_latest.xlsx')

generate_excel1('TLT_Policy_Compliance_Report_Windows_2016_Member_Server_Monthly_temp2.csv', 'TLT_Policy_Compliance_Report_Windows_2016_Member_Server_Monthly_latest.csv', 'TLT_Policy_Compliance_Report_Windows_2016_Member_Server_Monthly_latest.xlsx')

generate_excel1('TLT_Policy_Compliance_Report_Windows_10_Monthly_temp2.csv', 'TLT_Policy_Compliance_Report_Windows_10_Monthly_latest.csv', 'TLT_Policy_Compliance_Report_Windows_10_Monthly_latest.xlsx')

generate_excel2('TLT_Policy_Compliance_Report_Windows_2019_Domain_Server_Monthly_temp2.csv','TLT_Policy_Compliance_Report_Windows_2019_Domain_Server_Monthly_ByHost_ByConTrol.xlsx')

generate_excel2('TLT_Policy_Compliance_Report_Windows_2019_Member_Server_Monthly_temp2.csv','TLT_Policy_Compliance_Report_Windows_2019_Member_Server_Monthly_ByHost_ByConTrol.xlsx')

generate_excel2('TLT_Policy_Compliance_Report_Windows_2016_Member_Server_Monthly_temp2.csv','TLT_Policy_Compliance_Report_Windows_2016_Member_Server_Monthly_ByHost_ByConTrol.xlsx')

generate_excel2('TLT_Policy_Compliance_Report_Windows_10_Monthly_temp2.csv','TLT_Policy_Compliance_Report_Windows_10_Monthly_ByHost_ByConTrol.xlsx')
